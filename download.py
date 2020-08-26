import openpyxl
import requests
import time
import os

class ExcelOp(object):
    def __init__(self, file):
        self.file = file
        self.wb = openpyxl.load_workbook(self.file)
        sheets = self.wb.sheetnames
        self.sheet = sheets[0]
        self.ws = self.wb[self.sheet]

    # 获取表格的总行数和总列数
    def get_row_clo_num(self):
        rows = self.ws.max_row
        columns = self.ws.max_column
        return rows, columns

    # 获取某个单元格的值
    def get_cell_value(self, row, column):
        cell_value = self.ws.cell(row=row, column=column).value
        return cell_value

    # 获取某列的所有值
    def get_col_value(self, column):
        rows = self.ws.max_row
        column_data = []
        for i in range(1, rows + 1):
            cell_value = self.ws.cell(row=i, column=column).value
            column_data.append(cell_value)
        return column_data

    # 获取某行所有值
    def get_row_value(self, row):
        columns = self.ws.max_column
        row_data = []
        for i in range(1, columns + 1):
            cell_value = self.ws.cell(row=row, column=i).value
            row_data.append(cell_value)
        return row_data

    # 设置某个单元格的值
    def set_cell_value(self, row, colunm, cellvalue):
        try:
            self.ws.cell(row=row, column=colunm).value = cellvalue
            self.wb.save(self.file)
        except:
            self.ws.cell(row=row, column=colunm).value = "writefail"
            self.wb.save(self.file)

class Downloader(object):

    def download_from_alldatasheet(self, filename, url, savefolder=""):
        SUCCESS = False
        filename = filename + ".pdf"
        # url = "https://pdf1.alldatasheet.com/datasheet-pdf/download/79516/INFINEON/BTS6163D.html"
        params = {
            "tmpinfo1aa": "abc",
        }
        while not SUCCESS:
            print("尝试URL连接中...")
            req = requests.request(method='POST', url=url, params=params)
            print("获取返回内容信息文本，如果是pdf，比较耗时，请等待...")
            text = req.text
            print("获取返回内容信息文本完成。")
            if text.find("Download is temporarily unavailable") != -1:
                print("下载太频繁，等待五分钟......，等待中")
                time.sleep(300)
                continue
            else:
                SUCCESS = True
            # print(filename + " : \n")
            # print(text)
            data = req.content
            full_savefolder = os.path.join(os.path.abspath(os.curdir), savefolder)
            if not os.path.exists(full_savefolder):
                os.mkdir(full_savefolder)
            full_filepath = os.path.join(full_savefolder, filename)
            print("尝试存储文件：" + filename)
            with open(full_filepath, "wb") as code:
                code.write(data)
                print("存储文件：" + filename + "完成")
                return True
        return False

    def valid_url(self, url):
        # https://www.alldatasheet.com/datasheet-pdf/pdf/436213/HITTITE/HMC704LP4E.html
        if str(url).startswith("https://www.alldatasheet.com/datasheet-pdf/pdf/"):
            return True
        return False

    def is_tbd(self, checkstr):
        if str(checkstr).strip().lower() == 'no':
            return True
        return False

    def transfer_url(self, url):
        # https://www.alldatasheet.com/datasheet-pdf/pdf/436213/HITTITE/HMC704LP4E.html
        # https://pdf1.alldatasheet.com/datasheet-pdf/download/436213/HITTITE/HMC704LP4E.html
        before = "https://www.alldatasheet.com/datasheet-pdf/pdf/"
        after = "https://pdf1.alldatasheet.com/datasheet-pdf/download/"
        return str(url).replace(before, after)

if __name__ == "__main__":
    excel_op = ExcelOp(file="LIST.xlsx")
    downloader = Downloader()
    # 1-''， 4-型号规格, 6-下载, 7-链接
    need_cols = [1,4,6,7]
    rows, columns = excel_op.get_row_clo_num()
    # 从第二行开始解析下载
    for i in range(2,rows+1):
        need_col_values = []
        for j in need_cols:
            cell_val = excel_op.get_cell_value(i,j)
            need_col_values.append(cell_val)

        print(need_col_values)

        savefolder = need_col_values[0]
        filename = need_col_values[1]
        str_TBD = need_col_values[2]
        url = need_col_values[3]

        if downloader.is_tbd(str_TBD) and downloader.valid_url(url):
            url = downloader.transfer_url(url)
            res = downloader.download_from_alldatasheet(filename, url, savefolder)
            if res:
                excel_op.set_cell_value(i, 6, 'Yes')